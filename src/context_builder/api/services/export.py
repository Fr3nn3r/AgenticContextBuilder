"""Export service for exporting workspace data to Excel."""

import json
import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

from context_builder.storage.filesystem import FileStorage

logger = logging.getLogger(__name__)


class ExportService:
    """Service for exporting workspace data to Excel format."""

    def __init__(self, storage: FileStorage):
        """Initialize the export service.

        Args:
            storage: FileStorage instance for the workspace.
        """
        self.storage = storage

    def export_to_excel(self, output_path: Path) -> Dict[str, int]:
        """Export all workspace data to an Excel file.

        Args:
            output_path: Path where the Excel file will be written.

        Returns:
            Dict with counts for each exported entity type.
        """
        wb = Workbook()
        # Remove the default sheet
        wb.remove(wb.active)

        stats = {}

        # Export each entity type to a separate sheet
        stats["claims"] = self._export_claims(wb)
        stats["documents"] = self._export_documents(wb)
        stats["runs"] = self._export_runs(wb)
        stats["claim_runs"] = self._export_claim_runs(wb)
        stats["extractions"] = self._export_extractions(wb)
        stats["claim_facts"] = self._export_claim_facts(wb)
        stats["labels"] = self._export_labels(wb)
        stats["reconciliation"] = self._export_reconciliation(wb)

        # Save workbook
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

        return stats

    def _style_header(self, ws, num_cols: int) -> None:
        """Apply header styling to the first row."""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        for col in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    def _auto_width(self, ws) -> None:
        """Auto-adjust column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            # Cap width at 50 characters
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _export_claims(self, wb: Workbook) -> int:
        """Export claims to a worksheet."""
        ws = wb.create_sheet("Claims")
        headers = ["claim_id", "claim_folder", "doc_count"]
        ws.append(headers)
        self._style_header(ws, len(headers))

        claims = self.storage.list_claims()
        for claim in claims:
            ws.append([claim.claim_id, claim.claim_folder, claim.doc_count])

        self._auto_width(ws)
        return len(claims)

    def _export_documents(self, wb: Workbook) -> int:
        """Export documents to a worksheet."""
        ws = wb.create_sheet("Documents")
        headers = [
            "doc_id", "claim_id", "claim_folder", "doc_type", "filename",
            "source_type", "language", "page_count", "has_pdf", "has_text"
        ]
        ws.append(headers)
        self._style_header(ws, len(headers))

        count = 0
        claims = self.storage.list_claims()
        for claim in claims:
            docs = self.storage.list_docs(claim.claim_id)
            for doc in docs:
                ws.append([
                    doc.doc_id,
                    doc.claim_id,
                    doc.claim_folder,
                    doc.doc_type,
                    doc.filename,
                    doc.source_type,
                    doc.language,
                    doc.page_count,
                    doc.has_pdf,
                    doc.has_text,
                ])
                count += 1

        self._auto_width(ws)
        return count

    def _export_runs(self, wb: Workbook) -> int:
        """Export runs to a worksheet."""
        ws = wb.create_sheet("Runs")
        headers = [
            "run_id", "status", "started_at", "ended_at",
            "claims_count", "docs_count"
        ]
        ws.append(headers)
        self._style_header(ws, len(headers))

        runs = self.storage.list_runs()
        for run in runs:
            ws.append([
                run.run_id,
                run.status,
                run.started_at,
                run.ended_at,
                run.claims_count,
                run.docs_count,
            ])

        self._auto_width(ws)
        return len(runs)

    def _export_claim_runs(self, wb: Workbook) -> int:
        """Export claim runs to a worksheet."""
        ws = wb.create_sheet("Claim_Runs")
        headers = [
            "claim_run_id", "claim_id", "created_at", "stages_completed",
            "extraction_runs_considered", "contextbuilder_version"
        ]
        ws.append(headers)
        self._style_header(ws, len(headers))

        count = 0
        claims = self.storage.list_claims()
        for claim in claims:
            try:
                claim_run_storage = self.storage.get_claim_run_storage(claim.claim_id)
                claim_run_ids = claim_run_storage.list_claim_runs()
                for claim_run_id in claim_run_ids:
                    manifest = claim_run_storage.read_manifest(claim_run_id)
                    if manifest:
                        ws.append([
                            claim_run_id,
                            claim.claim_id,
                            manifest.created_at.isoformat() if manifest.created_at else None,
                            ", ".join(manifest.stages_completed),
                            ", ".join(manifest.extraction_runs_considered),
                            manifest.contextbuilder_version,
                        ])
                        count += 1
            except (ValueError, Exception) as e:
                logger.debug(f"No claim runs for {claim.claim_id}: {e}")

        self._auto_width(ws)
        return count

    def _export_extractions(self, wb: Workbook) -> int:
        """Export extractions to a worksheet (pivot format: one row per field)."""
        ws = wb.create_sheet("Extractions")
        headers = [
            "run_id", "doc_id", "claim_id", "doc_type",
            "field_name", "value", "confidence", "status",
            "has_verified_evidence"
        ]
        ws.append(headers)
        self._style_header(ws, len(headers))

        count = 0
        runs = self.storage.list_runs()

        for run in runs:
            extractions = self.storage.list_extractions(run.run_id)
            for ext_ref in extractions:
                extraction_data = self.storage.get_extraction(
                    run.run_id, ext_ref.doc_id, ext_ref.claim_id
                )
                if not extraction_data:
                    continue

                # Get doc type from extraction
                doc_info = extraction_data.get("doc", {})
                doc_type = doc_info.get("doc_type", "unknown")

                # Extract fields
                fields = extraction_data.get("fields", [])
                for field in fields:
                    value = field.get("value")
                    # Convert list values to string
                    if isinstance(value, list):
                        value = "; ".join(str(v) for v in value)

                    ws.append([
                        run.run_id,
                        ext_ref.doc_id,
                        ext_ref.claim_id,
                        doc_type,
                        field.get("name", ""),
                        value,
                        field.get("confidence", 0.0),
                        field.get("status", ""),
                        field.get("has_verified_evidence", False),
                    ])
                    count += 1

        self._auto_width(ws)
        return count

    def _export_claim_facts(self, wb: Workbook) -> int:
        """Export claim facts to a worksheet (pivot format: one row per fact)."""
        ws = wb.create_sheet("Claim_Facts")
        headers = [
            "claim_run_id", "claim_id", "fact_name", "value",
            "confidence", "source_doc_id", "source_doc_type", "source_run_id"
        ]
        ws.append(headers)
        self._style_header(ws, len(headers))

        count = 0
        claims = self.storage.list_claims()

        for claim in claims:
            try:
                claim_run_storage = self.storage.get_claim_run_storage(claim.claim_id)
                claim_run_ids = claim_run_storage.list_claim_runs()

                for claim_run_id in claim_run_ids:
                    facts_data = claim_run_storage.read_from_claim_run(claim_run_id, "facts.json")
                    if not facts_data:
                        continue

                    # Get facts from data
                    facts = facts_data.get("facts", [])
                    for fact in facts:
                        value = fact.get("value")
                        if isinstance(value, list):
                            value = "; ".join(str(v) for v in value)

                        selected_from = fact.get("selected_from", {})
                        ws.append([
                            claim_run_id,
                            claim.claim_id,
                            fact.get("name", ""),
                            value,
                            fact.get("confidence", 0.0),
                            selected_from.get("doc_id", ""),
                            selected_from.get("doc_type", ""),
                            selected_from.get("extraction_run_id", selected_from.get("run_id", "")),
                        ])
                        count += 1
            except (ValueError, Exception) as e:
                logger.debug(f"No claim facts for {claim.claim_id}: {e}")

        self._auto_width(ws)
        return count

    def _export_labels(self, wb: Workbook) -> int:
        """Export labels to a worksheet (pivot format: one row per field label)."""
        ws = wb.create_sheet("Labels")
        headers = [
            "doc_id", "claim_id", "field_name", "truth_value",
            "state", "unverifiable_reason", "reviewer", "reviewed_at"
        ]
        ws.append(headers)
        self._style_header(ws, len(headers))

        count = 0
        claims = self.storage.list_claims()

        for claim in claims:
            docs = self.storage.list_docs(claim.claim_id)
            for doc in docs:
                label_data = self.storage.get_label(doc.doc_id)
                if not label_data:
                    continue

                claim_id = label_data.get("claim_id", claim.claim_id)
                review = label_data.get("review", {})
                reviewer = review.get("reviewer", "")
                reviewed_at = review.get("reviewed_at", "")

                field_labels = label_data.get("field_labels", [])
                for fl in field_labels:
                    ws.append([
                        doc.doc_id,
                        claim_id,
                        fl.get("field_name", ""),
                        fl.get("truth_value"),
                        fl.get("state", ""),
                        fl.get("unverifiable_reason"),
                        reviewer,
                        reviewed_at,
                    ])
                    count += 1

        self._auto_width(ws)
        return count

    def _export_reconciliation(self, wb: Workbook) -> int:
        """Export reconciliation reports to a worksheet."""
        ws = wb.create_sheet("Reconciliation")
        headers = [
            "claim_id", "claim_run_id", "run_id", "gate_status",
            "fact_count", "conflict_count", "missing_critical_count",
            "missing_critical_facts", "provenance_coverage", "reasons"
        ]
        ws.append(headers)
        self._style_header(ws, len(headers))

        count = 0
        claims = self.storage.list_claims()

        for claim in claims:
            try:
                claim_run_storage = self.storage.get_claim_run_storage(claim.claim_id)
                claim_run_ids = claim_run_storage.list_claim_runs()

                for claim_run_id in claim_run_ids:
                    recon_data = claim_run_storage.read_from_claim_run(claim_run_id, "reconciliation.json")
                    if not recon_data:
                        continue

                    gate = recon_data.get("gate", {})
                    missing_critical = gate.get("missing_critical_facts", [])

                    ws.append([
                        claim.claim_id,
                        recon_data.get("claim_run_id", claim_run_id),
                        recon_data.get("run_id", ""),
                        gate.get("status", ""),
                        recon_data.get("fact_count", 0),
                        gate.get("conflict_count", 0),
                        len(missing_critical),
                        ", ".join(missing_critical) if missing_critical else "",
                        gate.get("provenance_coverage", 0.0),
                        ", ".join(gate.get("reasons", [])),
                    ])
                    count += 1
            except (ValueError, Exception) as e:
                logger.debug(f"No reconciliation for {claim.claim_id}: {e}")

        self._auto_width(ws)
        return count
